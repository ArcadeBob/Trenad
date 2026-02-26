# tests/test_excel_export.py
import os
import tempfile
from openpyxl import load_workbook
from pattern_scanner import PatternResult
from excel_export import export_to_excel


def _sample_results() -> list[PatternResult]:
    return [
        PatternResult(
            ticker="AAPL", pattern_type="Cup & Handle", confidence_score=85.0,
            buy_point=195.0, current_price=193.0, distance_to_pivot=-1.0,
            base_depth=22.0, base_length_weeks=14, volume_confirmation=True,
            above_50ma=True, above_200ma=True, rs_rating=88.0,
            pattern_details={"cup_low": 160.0},
        ),
        PatternResult(
            ticker="NVDA", pattern_type="Flat Base", confidence_score=72.0,
            buy_point=500.0, current_price=485.0, distance_to_pivot=-3.0,
            base_depth=9.0, base_length_weeks=6, volume_confirmation=False,
            above_50ma=True, above_200ma=True, rs_rating=95.0,
            pattern_details={},
        ),
    ]


def test_export_creates_file():
    results = _sample_results()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_report.xlsx")
        export_to_excel(results, path)
        assert os.path.exists(path)


def test_export_has_three_sheets():
    results = _sample_results()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_report.xlsx")
        export_to_excel(results, path)
        wb = load_workbook(path)
        assert "Pattern Scanner Results" in wb.sheetnames
        assert "Pattern Guide" in wb.sheetnames
        assert "Top Picks" in wb.sheetnames


def test_export_results_sheet_has_data():
    results = _sample_results()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_report.xlsx")
        export_to_excel(results, path)
        wb = load_workbook(path)
        ws = wb["Pattern Scanner Results"]
        # Header row + 2 data rows
        assert ws.max_row >= 3
        assert ws["A2"].value == "AAPL"
