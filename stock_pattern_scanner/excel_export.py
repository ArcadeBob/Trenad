"""Excel report generation for pattern scanner results."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pattern_scanner import PatternResult


def export_to_excel(results: list[PatternResult], filepath: str):
    """Export scan results to a formatted Excel workbook.

    Creates three sheets:
    - Pattern Scanner Results: full data table with conditional formatting
    - Pattern Guide: definitions of each pattern type
    - Top Picks: filtered actionable stocks
    """
    wb = Workbook()

    _create_results_sheet(wb, results)
    _create_guide_sheet(wb)
    _create_top_picks_sheet(wb, results)

    wb.save(filepath)


def _create_results_sheet(wb: Workbook, results: list[PatternResult]):
    ws = wb.active
    ws.title = "Pattern Scanner Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Ticker", "Pattern", "Score", "Status", "Price", "Buy Point",
        "Distance %", "Depth %", "Length (wks)", "RS Rating",
        "Above 50MA", "Above 200MA", "Vol Confirm",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        values = [
            r.ticker, r.pattern_type, r.confidence_score, r.status,
            r.current_price, r.buy_point,
            round(r.distance_to_pivot, 1), round(r.base_depth, 1),
            r.base_length_weeks, round(r.rs_rating, 1),
            "Yes" if r.above_50ma else "No",
            "Yes" if r.above_200ma else "No",
            "Yes" if r.volume_confirmation else "No",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Conditional formatting on score
        score_cell = ws.cell(row=row_idx, column=3)
        if r.confidence_score >= 75:
            score_cell.fill = green_fill
        elif r.confidence_score >= 60:
            score_cell.fill = yellow_fill

    # Column widths
    widths = [8, 18, 8, 12, 10, 10, 10, 10, 12, 10, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def _create_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("Pattern Guide")

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)

    ws.cell(row=1, column=1, value="Pattern Identification Guide").font = title_font

    patterns = [
        ("Cup & Handle", [
            "U-shaped base followed by small downward-drifting handle",
            "Cup depth: 12-33% correction from prior high",
            "Handle: 1-6 weeks, <15% decline, declining volume",
            "Duration: 7-65 weeks total",
            "Buy point: Break above handle high on volume surge",
        ]),
        ("Deep Cup & Handle", [
            "Same as Cup & Handle but 33-50% cup depth",
            "Often forms in volatile or bear markets",
            "Longer formation time typical",
        ]),
        ("Double Bottom", [
            "W-pattern with two distinct lows at similar price levels",
            "Lows within 3-5% of each other",
            "Second low may slightly undercut first (bullish shakeout)",
            "Depth: 20-30% typical correction",
            "Buy point: Break above middle peak",
        ]),
        ("Flat Base", [
            "Tight sideways consolidation, <15% range",
            "Minimum 5 weeks duration",
            "Often forms after breakout from deeper base",
            "Should form mostly above 50-day moving average",
            "Buy point: Break above base high",
        ]),
    ]

    row = 3
    for name, details in patterns:
        ws.cell(row=row, column=1, value=name).font = header_font
        row += 1
        for detail in details:
            ws.cell(row=row, column=1, value=f"  - {detail}")
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 70


def _create_top_picks_sheet(wb: Workbook, results: list[PatternResult]):
    ws = wb.create_sheet("Top Picks")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Filter: score >= 65, within 5% of buy point, above 50MA
    top_picks = [
        r for r in results
        if r.confidence_score >= 65
        and r.distance_to_pivot >= -5.0
        and r.above_50ma
    ]

    headers = ["Rank", "Ticker", "Pattern", "Score", "Status", "Price", "Buy Point", "Action"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(top_picks, 2):
        if r.status == "At Pivot":
            action = "WATCH - Near breakout"
        elif r.status == "Near Pivot":
            action = "WATCHLIST - Approaching"
        else:
            action = "MONITOR"

        values = [
            row_idx - 1, r.ticker, r.pattern_type, r.confidence_score,
            r.status, r.current_price, r.buy_point, action,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    widths = [6, 8, 18, 8, 12, 10, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
